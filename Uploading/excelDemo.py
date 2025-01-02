import os
import openpyxl

excelPath="C:/Users/psaik/Downloads/downloadExcel.xlsx"
sheet=openpyxl.load_workbook(excelPath)
ApplePrice=0

active=sheet.active
active.cell(row=3,column=4).value=400

Dict={}
row_column_Price=0

#For finding the row and column of price

for i in range(1,active.max_column+1):
    if active.cell(row=1,column=i).value=="price":
        row_column_Price=i
        print(f"Price lies in row 1 and {row_column_Price} column")
print(row_column_Price)

for i in range(1,active.max_row+1):
    
    for j in range(1,active.max_column+1):
            
        Dict[active.cell(row=i,column=2).value]= active.cell(row=i,column=row_column_Price).value
        

for key,value in Dict.items():
 
     if key=="Apple":
         ApplePrice=value
         print(f"The price of apple is {value}")

print(f"Again the apple price is {ApplePrice}")
       
sheet.save(excelPath)

print("Values Saved Successfully")

### This is the additional command I have added for checking git working ###

'''
for key,value in Dict.items():
     print(f"{key}: {value}")

'''
