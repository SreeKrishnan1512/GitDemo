import openpyxl

book=openpyxl.load_workbook("C:\\Users\\psaik\\Desktop\\Python_Selenium_Framework2\\TestData2\\pyExcelDemo.xlsx")
sheet=book.active
#value=sheet.cell(row=4,column=2).value

List=[]

for i in range(2,sheet.max_row+1):
    Dict={}
  
    for j in range(2,sheet.max_column+1):

        Dict[sheet.cell(row=1,column=j).value]= sheet.cell(row=i,column=j).value
        
        
    List.append(Dict)

print(List)





