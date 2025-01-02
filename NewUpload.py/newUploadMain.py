
import pytest
import openpyxl
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service



def Update_ExcelData(filePath,searchTerm,colName,new_value):

    Dict={}
    excelPath=filePath
    sheet=openpyxl.load_workbook(excelPath)
    active=sheet.active
    if not os.path.exists(excelPath):
        print("File not found!")
    else:
        print("File exists!")

    #To find price column
    for i in range(1,active.max_column+1):
        if active.cell(row=1,column=i).value==colName:
            Dict["Column"]=i

    #To find apple row
    for i in range(1,active.max_row+1):
        
        for j in range(1,active.max_column+1):
                
            if active.cell(row=i,column=j).value== searchTerm:
                Dict["Row"]=i
    

    print(Dict)

    active.cell(row=Dict["Row"],column=Dict["Column"]).value=new_value
    newValue=active.cell(row=Dict["Row"],column=Dict["Column"]).value
    print(newValue)
        
    sheet.save(excelPath)
    print("Values Saved Successfully")


chromePath="C:/Users/psaik/Desktop/Selenium_Excel_New/chromedriver-win64/chromedriver-win64/chromedriver.exe"
excelPath="C:/Users/psaik/Downloads/downloadExcel.xlsx"
service_obj= Service(chromePath)     
driver=webdriver.Chrome(service=service_obj)


driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.maximize_window()

fruit_name="Apple"
newValue=990

WebDriverWait(driver,10).until(
    EC.visibility_of_element_located((By.CSS_SELECTOR,"#downloadButton"))
    ).click()

Update_ExcelData(excelPath,fruit_name,"price",newValue)

input_file=WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"input[type='file']")))



input_file.send_keys(excelPath)


text_capture=WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH,"//div[text()='Updated Excel Data Successfully.']")))
txt=text_capture.text
priceColumn= driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price=driver.find_element(By.XPATH,f"//div[text()='{fruit_name}']/parent::div/parent::div/div[@id='cell-{priceColumn}-undefined']").text
print(actual_price)
print(txt)
#time.sleep(5)

assert actual_price== str(newValue),"Value mismatch"

print("passed")

