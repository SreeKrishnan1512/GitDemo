import os
import openpyxl

excelPath="C:/Users/psaik/Downloads/downloadExcel.xlsx"
sheet=openpyxl.load_workbook(excelPath)
ApplePrice=0

active=sheet.active
#active.cell(row=3,column=4).value=400

Dict={}

#To find price column

for i in range(1,active.max_column+1):
    if active.cell(row=1,column=i).value=="price":
        Dict["priceColumn"]=i

#To find apple row
for i in range(1,active.max_row+1):
    
    for j in range(1,active.max_column+1):
            
        if active.cell(row=i,column=j).value== "Apple":
            Dict["AppleRow"]=i

print(Dict)

active.cell(row=Dict["AppleRow"],column=Dict["priceColumn"]).value=1100
newValue=active.cell(row=Dict["AppleRow"],column=Dict["priceColumn"])
print(newValue.value)
       
sheet.save(excelPath)
print("Values Saved Successfully")